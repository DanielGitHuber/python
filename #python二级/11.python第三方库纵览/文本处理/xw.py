from openpyxl import *
# 读单元
wb = load_workbook('sample_2.xlsx')
ws = wb.active

for i in range(1, 13):
    for j in range(1, 7):
        print(ws.cell(row=i, column=j).value)

cells = ws['A1':'B12']
for c1, c2 in cells:
    print('{0:<15}\t{1:<15}'.format(c1.value, c2.value))


